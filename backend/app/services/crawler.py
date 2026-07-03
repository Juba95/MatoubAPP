"""
Crawler SEO (mini Screaming Frog) — 100 % gratuit (httpx + BeautifulSoup).

Crawle un site en restant sur le domaine, extrait pour chaque page les
signaux SEO (status, title, meta description, H1, Hn, canonical, robots,
liens internes/externes, nombre de mots), détecte la ville d'une page à
partir de la base de communes, vérifie la cohérence Title/H1/ville, mesure
la proximité de contenu entre pages, puis analyse le maillage interne
(pages orphelines, inlinks, ancres) avec des suggestions adaptées à une
stratégie « site service + géolocalisé ».
"""
import csv
import os
import re
import unicodedata
from collections import defaultdict, deque
from concurrent.futures import ThreadPoolExecutor, as_completed
from urllib.parse import urljoin, urlparse, urldefrag

import httpx
from bs4 import BeautifulSoup

USER_AGENT = (
    "Mozilla/5.0 (compatible; MatoubBot/1.0; +https://matoub.app) "
    "SEO-Crawler"
)

_ASSET_RE = re.compile(r"\.(jpe?g|png|gif|svg|webp|ico|css|js|pdf|zip|mp4|mp3|woff2?|ttf|xml|json)(\?|$)", re.I)
_WORD_RE = re.compile(r"[a-zà-ÿ0-9]+", re.I)


def _norm(text: str) -> str:
    """Minuscule ASCII sans accents."""
    text = unicodedata.normalize("NFKD", text or "")
    text = text.encode("ascii", "ignore").decode("ascii")
    return text.lower().strip()


# ---------------------------------------------------------------------------
# Base de villes (détection de ville dans les pages)
# ---------------------------------------------------------------------------

_CITY_CACHE: dict | None = None


def _villes_csv_path() -> str | None:
    candidates = [
        os.path.join(os.path.dirname(__file__), "..", "..", "..", "data", "villes_france_premium.csv"),
        os.path.join(os.path.dirname(__file__), "..", "..", "data", "villes_france_premium.csv"),
        "/app/data/villes_france_premium.csv",
        os.path.join(os.path.dirname(__file__), "..", "..", "villes_france.csv"),
    ]
    return next((p for p in candidates if os.path.exists(p)), None)


def _load_cities() -> dict:
    """Retourne {slug_norm: NomVille}. Ignore les villes au nom trop court/commun."""
    global _CITY_CACHE
    if _CITY_CACHE is not None:
        return _CITY_CACHE
    cities: dict[str, str] = {}
    path = _villes_csv_path()
    if path:
        try:
            with open(path, encoding="utf-8") as f:
                for row in csv.reader(f):
                    try:
                        display = row[5]           # "Lyon", "Saint-Étienne"
                        slug = _norm(row[2])       # "lyon", "saint-etienne"
                    except IndexError:
                        continue
                    if len(slug) < 4:              # évite "y", "eu"... faux positifs
                        continue
                    cities.setdefault(slug, display)
        except Exception:
            pass
    _CITY_CACHE = cities
    return cities


def detect_city(url: str, h1: str, title: str) -> str:
    """Détecte la ville d'une page (priorité : slug d'URL, puis H1, puis title)."""
    cities = _load_cities()
    if not cities:
        return ""
    # 1. Depuis le slug d'URL (le plus fiable pour les pages géoloc)
    path = _norm(urlparse(url).path)
    path = re.sub(r"\.(html?|php|aspx?)$", "", path)   # retire l'extension éventuelle
    tokens = [t for t in re.split(r"[-/_]", path) if t]
    # bigrammes puis unigrammes (pour "saint-etienne")
    for size in (3, 2, 1):
        for i in range(len(tokens) - size + 1):
            cand = "-".join(tokens[i:i + size])
            if cand in cities:
                return cities[cand]
    # 2. Depuis H1 puis title
    for text in (h1, title):
        norm = _norm(text)
        words = [w for w in re.split(r"[\s,|/–—-]+", norm) if len(w) >= 4]
        for size in (3, 2, 1):
            for i in range(len(words) - size + 1):
                cand = "-".join(words[i:i + size])
                if cand in cities:
                    return cities[cand]
    return ""


def _shingles(text: str, n: int = 5) -> frozenset:
    words = _WORD_RE.findall(text.lower())
    return frozenset(
        hash(" ".join(words[i:i + n])) for i in range(max(0, len(words) - n + 1))
    )


# Mots très fréquents par langue (heuristique de détection, gratuite)
_LANG_STOPWORDS = {
    "fr": {"le", "la", "les", "de", "des", "un", "une", "et", "est", "vous", "nous",
           "pour", "dans", "avec", "sur", "que", "qui", "notre", "votre", "plus", "à"},
    "en": {"the", "and", "of", "to", "in", "for", "you", "we", "our", "your", "with",
           "is", "are", "this", "that", "on", "at", "by", "from", "more"},
    "es": {"el", "la", "los", "las", "de", "un", "una", "y", "es", "para", "con",
           "que", "su", "nuestro", "por", "más", "en", "se", "como", "del"},
    "de": {"der", "die", "das", "und", "ist", "für", "sie", "wir", "unser", "mit",
           "von", "auf", "den", "im", "ein", "eine", "zu", "auch", "als", "mehr"},
    "it": {"il", "la", "di", "un", "una", "e", "che", "per", "con", "non", "sono",
           "nostro", "vostro", "più", "come", "questo", "in", "del", "dei", "gli"},
    "pt": {"o", "a", "os", "as", "de", "um", "uma", "e", "que", "para", "com", "não",
           "nosso", "por", "mais", "em", "do", "da", "como", "seu"},
    "nl": {"de", "het", "een", "van", "en", "is", "voor", "met", "op", "wij", "onze",
           "uw", "dat", "die", "aan", "te", "in", "zijn", "meer", "ook"},
}


def detect_language(html_lang: str, text: str) -> str:
    """Détecte la langue d'une page : attribut html lang en priorité, sinon
    heuristique par mots fréquents (gratuit, aucune dépendance)."""
    if html_lang:
        code = re.split(r"[-_]", html_lang.strip().lower())[0]
        if code in _LANG_STOPWORDS or len(code) == 2:
            return code
    words = _WORD_RE.findall(text.lower())[:1500]
    if len(words) < 20:
        return ""
    wset = set(words)
    scores = {lg: sum(1 for w in words if w in sw) for lg, sw in _LANG_STOPWORDS.items()}
    best = max(scores, key=scores.get)
    # Nécessite un minimum de signal pour ne pas deviner au hasard
    return best if scores[best] >= max(3, len(words) * 0.02) else ""


# ---------------------------------------------------------------------------
# Crawler
# ---------------------------------------------------------------------------


class SiteCrawler:
    """Crawle un site et produit un rapport SEO par page + analyse de maillage."""

    # Plafond dur (garde-fou mémoire/temps). Au-delà, préférer un vrai crawler dédié.
    MAX_PAGES_CAP = 50000
    MAX_WORKERS_CAP = 32
    # Au-delà de ce nombre de pages, la détection de quasi-doublons (O(n²)) est
    # désactivée par défaut (trop coûteuse) sauf si explicitement demandée.
    PROXIMITY_AUTO_LIMIT = 3000

    def __init__(
        self,
        start_url: str,
        max_pages: int = 300,
        max_workers: int = 0,
        include_sitemap: bool = True,
        sitemap_url: str = "",
        timeout: float = 15.0,
        fast: bool = False,
        compute_proximity: bool | None = None,
        progress=None,
    ):
        if not start_url.startswith(("http://", "https://")):
            start_url = "https://" + start_url
        self.start_url = start_url.rstrip("/")
        parsed = urlparse(self.start_url)
        self.scheme = parsed.scheme
        self.host = parsed.netloc.lower()
        self.root_host = self.host[4:] if self.host.startswith("www.") else self.host
        self.max_pages = max(1, min(max_pages, self.MAX_PAGES_CAP))
        self.fast = fast
        # Auto-scaling des workers : plus de pages → plus de parallélisme.
        if not max_workers or max_workers <= 0:
            if self.max_pages > 10000 or fast:
                max_workers = 32
            elif self.max_pages > 2000:
                max_workers = 24
            elif self.max_pages > 500:
                max_workers = 16
            else:
                max_workers = 8
        self.max_workers = max(1, min(max_workers, self.MAX_WORKERS_CAP))
        self.include_sitemap = include_sitemap
        self.sitemap_url = sitemap_url.strip()
        self.timeout = timeout
        # Proximité : auto (désactivée sur gros crawls ou en mode rapide) sauf override.
        if compute_proximity is None:
            compute_proximity = (not fast) and (self.max_pages <= self.PROXIMITY_AUTO_LIMIT)
        self.compute_proximity = compute_proximity
        self.progress = progress or (lambda done, total: None)

        self.pages: dict[str, dict] = {}      # url -> page data
        self._seen: set[str] = set()
        self._queue: deque[str] = deque()

    # -- helpers URL -------------------------------------------------------

    def _same_site(self, url: str) -> bool:
        h = urlparse(url).netloc.lower()
        h = h[4:] if h.startswith("www.") else h
        return h == self.root_host

    def _clean(self, url: str) -> str:
        url, _ = urldefrag(url)          # retire #ancre
        return url.rstrip("/") or url

    # -- sitemap -----------------------------------------------------------

    def _discover_sitemap_urls(self, client: httpx.Client) -> list[str]:
        urls: list[str] = []
        candidates = [self.sitemap_url] if self.sitemap_url else [
            f"{self.scheme}://{self.host}/sitemap.xml",
            f"{self.scheme}://{self.host}/sitemap_index.xml",
        ]
        # robots.txt peut pointer un sitemap
        try:
            r = client.get(f"{self.scheme}://{self.host}/robots.txt")
            if r.status_code == 200:
                for m in re.findall(r"(?i)sitemap:\s*(\S+)", r.text):
                    candidates.append(m.strip())
        except Exception:
            pass

        to_fetch = list(dict.fromkeys(c for c in candidates if c))
        seen_sm: set[str] = set()
        while to_fetch and len(urls) < self.max_pages * 2:
            sm = to_fetch.pop(0)
            if sm in seen_sm:
                continue
            seen_sm.add(sm)
            try:
                r = client.get(sm)
                if r.status_code != 200:
                    continue
                locs = re.findall(r"<loc>\s*(.*?)\s*</loc>", r.text, re.I | re.S)
                if "<sitemapindex" in r.text.lower():
                    to_fetch.extend(loc.strip() for loc in locs)   # index → sous-sitemaps
                else:
                    urls.extend(loc.strip() for loc in locs if self._same_site(loc.strip()))
            except Exception:
                continue
        return list(dict.fromkeys(urls))

    # -- fetch + parse -----------------------------------------------------

    def _fetch(self, client: httpx.Client, url: str) -> dict:
        data = {
            "url": url, "status": 0, "title": "", "title_len": 0,
            "meta_description": "", "desc_len": 0, "h1": "", "h1_count": 0,
            "h2_count": 0, "h3_count": 0, "hn": [], "word_count": 0,
            "canonical": "", "meta_robots": "", "indexable": True,
            "internal_links": [], "external_links": 0, "anchors_out": [],
            "content_type": "", "error": "",
            "html_lang": "", "language": "", "hreflangs": [], "hreflang_count": 0,
        }
        try:
            r = client.get(url)
        except Exception as exc:
            data["error"] = str(exc)[:200]
            return data
        data["status"] = r.status_code
        data["content_type"] = r.headers.get("content-type", "").split(";")[0]
        if r.status_code >= 400 or "text/html" not in data["content_type"]:
            data["indexable"] = False
            return data

        soup = BeautifulSoup(r.text, "html.parser")

        # Langue déclarée (<html lang="...">) + versions hreflang
        html_tag = soup.find("html")
        data["html_lang"] = (html_tag.get("lang", "").strip() if html_tag else "")
        hreflangs = []
        for link in soup.find_all("link", attrs={"rel": re.compile(r"alternate", re.I)}):
            hl = (link.get("hreflang") or "").strip()
            if hl:
                hreflangs.append({"hreflang": hl.lower(), "href": urljoin(url, link.get("href", ""))})
        data["hreflangs"] = hreflangs
        data["hreflang_count"] = len(hreflangs)

        title_tag = soup.find("title")
        data["title"] = title_tag.get_text(strip=True) if title_tag else ""
        data["title_len"] = len(data["title"])

        md = soup.find("meta", attrs={"name": re.compile(r"^description$", re.I)})
        if md and md.get("content"):
            data["meta_description"] = md["content"].strip()
            data["desc_len"] = len(data["meta_description"])

        h1s = soup.find_all("h1")
        data["h1"] = h1s[0].get_text(strip=True) if h1s else ""
        data["h1_count"] = len(h1s)
        data["h2_count"] = len(soup.find_all("h2"))
        data["h3_count"] = len(soup.find_all("h3"))
        data["hn"] = [
            {"level": int(h.name[1]), "text": h.get_text(strip=True)[:120]}
            for h in soup.find_all(["h1", "h2", "h3", "h4"])
            if h.get_text(strip=True)
        ][:60]

        can = soup.find("link", attrs={"rel": re.compile(r"canonical", re.I)})
        if can and can.get("href"):
            data["canonical"] = urljoin(url, can["href"].strip())

        mr = soup.find("meta", attrs={"name": re.compile(r"^robots$", re.I)})
        if mr and mr.get("content"):
            data["meta_robots"] = mr["content"].strip().lower()
            if "noindex" in data["meta_robots"]:
                data["indexable"] = False

        # Contenu texte (retire nav/footer/script/style pour le word count)
        body = BeautifulSoup(r.text, "html.parser")
        for tag in body(["script", "style", "nav", "footer", "header", "aside", "noscript"]):
            tag.decompose()
        text = body.get_text(" ", strip=True)
        data["word_count"] = len(_WORD_RE.findall(text))
        data["_text"] = text[:20000]     # pour la proximité (non exposé)
        data["language"] = detect_language(data["html_lang"], text)

        internal, external = [], 0
        anchors = []
        for a in soup.find_all("a", href=True):
            href = a["href"].strip()
            if href.startswith(("mailto:", "tel:", "javascript:", "#")):
                continue
            full = self._clean(urljoin(url, href))
            if not full.startswith(("http://", "https://")):
                continue
            if _ASSET_RE.search(full):
                continue
            if self._same_site(full):
                internal.append(full)
                anchors.append({"to": full, "anchor": a.get_text(strip=True)[:80],
                                "rel": a.get("rel", [])})
            else:
                external += 1
        data["internal_links"] = list(dict.fromkeys(internal))
        data["external_links"] = external
        data["anchors_out"] = anchors

        # Détection ville + cohérences
        city = detect_city(url, data["h1"], data["title"])
        data["city"] = city
        data["title_h1_match"] = _title_h1_similar(data["title"], data["h1"])
        data["city_in_title"] = bool(city) and _norm(city) in _norm(data["title"])
        data["city_in_h1"] = bool(city) and _norm(city) in _norm(data["h1"])
        return data

    # -- boucle de crawl ---------------------------------------------------

    def run(self) -> dict:
        headers = {"User-Agent": USER_AGENT, "Accept-Language": "fr-FR,fr;q=0.9"}
        with httpx.Client(follow_redirects=True, timeout=self.timeout, headers=headers) as client:
            # Amorçage
            seeds = [self.start_url]
            if self.include_sitemap:
                seeds += self._discover_sitemap_urls(client)
            for s in seeds:
                cs = self._clean(s)
                if cs not in self._seen and self._same_site(cs):
                    self._seen.add(cs)
                    self._queue.append(cs)

            # BFS parallélisé par vagues
            with ThreadPoolExecutor(max_workers=self.max_workers) as pool:
                while self._queue and len(self.pages) < self.max_pages:
                    batch = []
                    while self._queue and len(batch) < self.max_workers and \
                            len(self.pages) + len(batch) < self.max_pages:
                        batch.append(self._queue.popleft())
                    futures = {pool.submit(self._fetch, client, u): u for u in batch}
                    for fut in as_completed(futures):
                        page = fut.result()
                        self.pages[page["url"]] = page
                        # Enfiler les nouveaux liens internes
                        for link in page.get("internal_links", []):
                            if link not in self._seen and len(self._seen) < self.max_pages * 3:
                                self._seen.add(link)
                                self._queue.append(link)
                    self.progress(len(self.pages), min(self.max_pages, len(self._seen)))

        if self.compute_proximity:
            self._compute_proximity()
        else:
            # Nettoie le texte non utilisé et neutralise les champs de proximité
            for p in self.pages.values():
                p.pop("_text", None)
                p["max_similarity"] = 0
                p["closest_page"] = ""
        return self._report()

    # -- proximité de contenu ---------------------------------------------

    def _compute_proximity(self):
        html_pages = [
            p for p in self.pages.values()
            if p.get("status", 0) < 400 and p.get("_text") and p.get("word_count", 0) > 50
        ]
        sample = html_pages[:300]   # borne l'O(n²)
        shingles = {p["url"]: _shingles(p.pop("_text")) for p in sample}
        # retire _text des pages non échantillonnées
        for p in self.pages.values():
            p.pop("_text", None)
        urls = list(shingles.keys())
        for p in self.pages.values():
            p["max_similarity"] = 0
            p["closest_page"] = ""
        for i, u1 in enumerate(urls):
            s1 = shingles[u1]
            if not s1:
                continue
            best, best_url = 0.0, ""
            for j, u2 in enumerate(urls):
                if i == j:
                    continue
                s2 = shingles[u2]
                if not s2:
                    continue
                inter = len(s1 & s2)
                if not inter:
                    continue
                jac = inter / len(s1 | s2)
                if jac > best:
                    best, best_url = jac, u2
            self.pages[u1]["max_similarity"] = round(best * 100)
            self.pages[u1]["closest_page"] = best_url

    # -- rapport final -----------------------------------------------------

    def _report(self) -> dict:
        from app.services.internal_linking import analyze_internal_linking

        pages = list(self.pages.values())
        # Audit hreflang (réciprocité + x-default) sur l'ensemble des pages
        hreflang_audit = _audit_hreflang(pages)
        for p in pages:
            p["internal_out"] = len(p.get("internal_links", []))
            p["hreflang_issue"] = hreflang_audit["by_url"].get(p["url"], "")
            p["issues"] = _page_issues(p)
        maillage = analyze_internal_linking(pages, self.start_url)
        # Injecte les inlinks calculés par le maillage dans chaque page
        inlinks = maillage.get("inlinks_by_url", {})
        for p in pages:
            p["inlinks"] = inlinks.get(p["url"], 0)

        html_pages = [p for p in pages if p.get("content_type") == "text/html"]
        summary = {
            "start_url": self.start_url,
            "total_pages": len(pages),
            "html_pages": len(html_pages),
            "errors_4xx_5xx": sum(1 for p in pages if p.get("status", 0) >= 400),
            "noindex": sum(1 for p in html_pages if not p.get("indexable", True)),
            "missing_title": sum(1 for p in html_pages if not p.get("title")),
            "missing_desc": sum(1 for p in html_pages if not p.get("meta_description")),
            "missing_h1": sum(1 for p in html_pages if not p.get("h1")),
            "multiple_h1": sum(1 for p in html_pages if p.get("h1_count", 0) > 1),
            "title_h1_mismatch": sum(1 for p in html_pages if p.get("h1") and not p.get("title_h1_match")),
            "duplicate_content": sum(1 for p in html_pages if p.get("max_similarity", 0) >= 90),
            "city_pages": sum(1 for p in html_pages if p.get("city")),
            "city_missing_in_title": sum(1 for p in html_pages if p.get("city") and not p.get("city_in_title")),
            "orphan_pages": len(maillage.get("orphans", [])),
            "unreachable_pages": len(maillage.get("unreachable", [])),
            "deep_pages": len(maillage.get("deep_pages", [])),
            "cannibalization_groups": len(maillage.get("cannibalization", [])),
            "maillage_health": maillage.get("health_score", {}).get("score"),
            "pages_with_issues": sum(1 for p in pages if p.get("issues")),
        }
        # Multilingue
        lang_counts: dict[str, int] = {}
        for p in html_pages:
            lg = p.get("language") or "?"
            lang_counts[lg] = lang_counts.get(lg, 0) + 1
        summary["languages"] = lang_counts
        summary["is_multilingual"] = len([l for l in lang_counts if l != "?"]) > 1
        summary["missing_html_lang"] = sum(1 for p in html_pages if not p.get("html_lang"))
        summary["pages_with_hreflang"] = sum(1 for p in html_pages if p.get("hreflang_count", 0) > 0)
        summary["hreflang_issues"] = hreflang_audit["issue_count"]
        summary["missing_hreflang_multilingual"] = (
            sum(1 for p in html_pages if p.get("hreflang_count", 0) == 0)
            if summary["is_multilingual"] else 0
        )
        return {"summary": summary, "pages": pages, "maillage": maillage}


def _audit_hreflang(pages: list[dict]) -> dict:
    """Vérifie la réciprocité des hreflang et la présence de x-default.

    Google exige que si A déclare une alternative B, B déclare aussi A.
    Retourne {by_url: {url: message}, issue_count}.
    """
    def _norm_url(u: str) -> str:
        return u.split("#")[0].rstrip("/")

    known = {_norm_url(p["url"]) for p in pages}
    # href déclarés par chaque page (hors x-default)
    declared: dict[str, set] = {}
    has_xdefault: dict[str, bool] = {}
    for p in pages:
        u = _norm_url(p["url"])
        hrefs, xdef = set(), False
        for h in p.get("hreflangs", []):
            if h["hreflang"] == "x-default":
                xdef = True
            else:
                hrefs.add(_norm_url(h["href"]))
        declared[u] = hrefs
        has_xdefault[u] = xdef

    by_url: dict[str, str] = {}
    for p in pages:
        u = _norm_url(p["url"])
        hrefs = declared.get(u, set())
        if not hrefs:
            continue
        problems = []
        if not has_xdefault[u]:
            problems.append("x-default absent")
        for target in hrefs:
            if target == u:
                continue
            if target in known and u not in declared.get(target, set()):
                problems.append(f"non réciproque ({target.rsplit('/', 1)[-1] or target})")
                break  # un exemple suffit
        if problems:
            by_url[p["url"]] = " ; ".join(problems)
    return {"by_url": by_url, "issue_count": len(by_url)}


def _title_h1_similar(title: str, h1: str) -> bool:
    """Le title et le H1 partagent-ils l'essentiel de leurs mots ?"""
    if not title or not h1:
        return False
    tw = set(w for w in _WORD_RE.findall(_norm(title)) if len(w) > 2)
    hw = set(w for w in _WORD_RE.findall(_norm(h1)) if len(w) > 2)
    if not tw or not hw:
        return False
    return len(tw & hw) / len(hw) >= 0.5


def _page_issues(p: dict) -> list[str]:
    """Liste des problèmes SEO d'une page (labels courts)."""
    issues = []
    st = p.get("status", 0)
    if st >= 400:
        issues.append(f"HTTP {st}")
        return issues
    if p.get("content_type") != "text/html":
        return issues
    if not p.get("title"):
        issues.append("Title manquant")
    elif p["title_len"] > 65:
        issues.append("Title trop long")
    elif p["title_len"] < 20:
        issues.append("Title trop court")
    if not p.get("meta_description"):
        issues.append("Meta description manquante")
    elif p["desc_len"] > 160:
        issues.append("Meta desc trop longue")
    if not p.get("h1"):
        issues.append("H1 manquant")
    if p.get("h1_count", 0) > 1:
        issues.append("H1 multiples")
    if p.get("h1") and not p.get("title_h1_match"):
        issues.append("Title ≠ H1")
    if not p.get("indexable", True):
        issues.append("Noindex")
    if p.get("word_count", 0) < 250 and p.get("content_type") == "text/html":
        issues.append("Contenu léger (<250 mots)")
    if p.get("max_similarity", 0) >= 90:
        issues.append(f"Quasi-doublon ({p['max_similarity']}%)")
    if p.get("city") and not p.get("city_in_title"):
        issues.append("Ville absente du Title")
    if p.get("city") and not p.get("city_in_h1"):
        issues.append("Ville absente du H1")
    if not p.get("html_lang"):
        issues.append("Attribut lang manquant")
    if p.get("hreflang_issue"):
        issues.append(f"hreflang : {p['hreflang_issue']}")
    return issues


# ---------------------------------------------------------------------------
# Export Excel
# ---------------------------------------------------------------------------


def build_crawl_excel(report: dict) -> bytes:
    import io
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment

    wb = openpyxl.Workbook()

    # Feuille 1 : pages (vue Screaming Frog)
    ws = wb.active
    ws.title = "Pages"
    headers = [
        "URL", "Status", "Type", "Indexable", "Langue", "lang HTML",
        "Versions hreflang", "Title", "Long. Title",
        "Meta description", "Long. desc", "H1", "Nb H1", "Nb H2", "Nb H3",
        "Mots", "Canonical", "Meta robots", "Ville détectée",
        "Ville dans Title", "Title=H1", "Liens internes (out)", "Inlinks",
        "PageRank (0-100)", "Profondeur (clics)",
        "Similarité max %", "Page la + proche", "Problèmes",
    ]
    _write_header(ws, headers)
    for p in report["pages"]:
        depth = p.get("click_depth", -1)
        ws.append([
            p["url"], p.get("status"), p.get("content_type"),
            "Oui" if p.get("indexable", True) else "Non",
            p.get("language", ""), p.get("html_lang", ""), p.get("hreflang_count", 0),
            p.get("title", ""), p.get("title_len", 0),
            p.get("meta_description", ""), p.get("desc_len", 0),
            p.get("h1", ""), p.get("h1_count", 0), p.get("h2_count", 0), p.get("h3_count", 0),
            p.get("word_count", 0), p.get("canonical", ""), p.get("meta_robots", ""),
            p.get("city", ""), "Oui" if p.get("city_in_title") else ("Non" if p.get("city") else ""),
            "Oui" if p.get("title_h1_match") else "Non",
            p.get("internal_out", 0), p.get("inlinks", 0),
            p.get("pagerank", 0), ("inatteignable" if depth == -1 else depth),
            p.get("max_similarity", 0), p.get("closest_page", ""),
            " | ".join(p.get("issues", [])),
        ])

    # Feuille 2 : maillage — orphelines, inatteignables, profondes, top PageRank
    m = report.get("maillage", {})
    ws2 = wb.create_sheet("Maillage")
    _write_header(ws2, ["Type", "URL", "PageRank / Inlinks", "Détail"])
    for o in m.get("orphans", []):
        ws2.append(["Orpheline", o, 0, "Aucun lien interne entrant"])
    for o in m.get("unreachable", []):
        ws2.append(["Inatteignable", o, 0, "Non atteignable depuis l'accueil"])
    for dp in m.get("deep_pages", []):
        ws2.append(["Profonde", dp["url"], "", f"{dp['depth']} clics depuis l'accueil"])
    for t in m.get("top_linked", []):
        ws2.append(["Top PageRank", t["url"], t.get("pagerank", 0), f"{t.get('inlinks', 0)} inlinks"])
    for s in m.get("sinks", []):
        ws2.append(["Puits (jus piégé)", s["url"], s.get("pagerank", 0), "Reçoit du jus mais ne redistribue pas"])

    # Feuille 3 : suggestions scorées par impact
    ws3 = wb.create_sheet("Suggestions")
    _write_header(ws3, ["Priorité", "Impact estimé", "De (page)", "Vers (page)", "Ancre suggérée", "Raison"])
    for s in m.get("suggestions", []):
        ws3.append([s.get("priority", ""), s.get("impact", ""), s.get("from", ""),
                    s.get("to", ""), s.get("anchor", ""), s.get("reason", "")])

    # Feuille 4 : cannibalisation
    ws4 = wb.create_sheet("Cannibalisation")
    _write_header(ws4, ["Type", "Clé", "Pages en concurrence"])
    for c in m.get("cannibalization", []):
        ws4.append([c.get("type", ""), c.get("key", ""), "  |  ".join(c.get("urls", []))])

    # Feuille 5 : silos
    ws5 = wb.create_sheet("Silos")
    _write_header(ws5, ["Cluster service", "Pages", "Densité", "Cohésion", "Hub de tête", "Couverture hub %"])
    for si in m.get("silos", []):
        ws5.append([si["cluster"], si["pages"], si["density"], si["cohesion"],
                    si.get("hub", ""), si.get("hub_coverage", 0)])

    # Largeurs
    for sheet in wb.worksheets:
        for col in sheet.columns:
            letter = col[0].column_letter
            sheet.column_dimensions[letter].width = 22

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _write_header(ws, headers):
    from openpyxl.styles import Font, PatternFill, Alignment
    font = Font(bold=True, color="FFFFFF", size=11)
    fill = PatternFill(start_color="2E86C1", end_color="2E86C1", fill_type="solid")
    for i, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=i, value=h)
        c.font = font
        c.fill = fill
        c.alignment = Alignment(horizontal="center")
    ws.freeze_panes = "A2"
