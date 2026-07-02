"""
Moteur de contenu geolocalisé pour pages SEO.

Genere des pages uniques par ville a partir d'un texte source,
avec variantes de contenu, maillage interne par proximite
geographique, analyse concurrentielle et export Excel
(compatible WP All Import / Divi Builder).
"""

import io
import json
import logging
import math
import re
import unicodedata
from datetime import datetime, timedelta
from typing import Optional

import anthropic
import httpx
from bs4 import BeautifulSoup
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill

from app.config import get_settings

logger = logging.getLogger(__name__)


# ---------------------------------------------------------------------------
# Fonctions utilitaires
# ---------------------------------------------------------------------------


def slugify(text: str) -> str:
    """Convertit un texte en slug URL (ASCII, minuscules, tirets)."""
    text = unicodedata.normalize("NFKD", text)
    text = text.encode("ascii", "ignore").decode("ascii")
    text = text.lower().strip()
    text = re.sub(r"[^\w\s-]", "", text)
    text = re.sub(r"[\s_]+", "-", text)
    text = re.sub(r"-+", "-", text)
    return text.strip("-")


def haversine(lat1: float, lon1: float, lat2: float, lon2: float) -> float:
    """Distance en kilometres entre deux points GPS (formule de Haversine)."""
    R = 6_371.0  # rayon moyen de la Terre en km
    rlat1, rlon1, rlat2, rlon2 = map(math.radians, [lat1, lon1, lat2, lon2])
    dlat = rlat2 - rlat1
    dlon = rlon2 - rlon1
    a = (
        math.sin(dlat / 2) ** 2
        + math.cos(rlat1) * math.cos(rlat2) * math.sin(dlon / 2) ** 2
    )
    return R * 2 * math.asin(math.sqrt(a))


def replace_ville(html: str, ville: str) -> str:
    """Remplace le placeholder ``__VILLE__`` (insensible a la casse)."""
    return re.sub(r"__VILLE__", ville, html, flags=re.IGNORECASE)


def parse_faq(html: str) -> list[tuple[str, str]]:
    """Extrait les paires question/réponse d'un bloc FAQ (H3 = question, <p> = réponse)."""
    if not html:
        return []
    pairs = []
    chunks = re.split(r"<h3[^>]*>", html)
    for chunk in chunks[1:]:
        m = re.match(r"(.*?)</h3>(.*)", chunk, re.DOTALL)
        if not m:
            continue
        question = re.sub(r"<[^>]+>", "", m.group(1)).strip()
        answer = re.sub(r"<[^>]+>", " ", m.group(2))
        answer = re.sub(r"\s+", " ", answer).strip()
        if question and answer:
            pairs.append((question, answer[:800]))
    return pairs


def build_jsonld(
    site_name: str,
    site_domain: str,
    keyword: str,
    ville: dict,
    faq_pairs: list[tuple[str, str]] | None = None,
    avis: list[dict] | None = None,
    breadcrumb: list[tuple[str, str]] | None = None,
) -> str:
    """Construit les blocs Schema.org JSON-LD d'une page géolocalisée.

    Génère LocalBusiness (+AggregateRating), FAQPage et BreadcrumbList —
    c'est ce qui déclenche les rich snippets (étoiles, FAQ dépliée) en SERP.
    """
    base = f"https://{site_domain.replace('https://', '').replace('http://', '').rstrip('/')}"
    schemas = []

    business: dict = {
        "@context": "https://schema.org",
        "@type": "LocalBusiness",
        "name": site_name,
        "url": base,
        "description": f"{keyword.capitalize()} à {ville.get('nom', '')}",
        "areaServed": {
            "@type": "City" if ville.get("type") != "Département" else "AdministrativeArea",
            "name": ville.get("nom", ""),
        },
    }
    if ville.get("postal_code"):
        business["address"] = {
            "@type": "PostalAddress",
            "postalCode": ville["postal_code"],
            "addressLocality": ville.get("nom", ""),
            "addressCountry": "FR",
        }
    if ville.get("lat") and ville.get("lon"):
        business["geo"] = {
            "@type": "GeoCoordinates",
            "latitude": ville["lat"],
            "longitude": ville["lon"],
        }
    if avis:
        notes = []
        for a in avis:
            try:
                notes.append(max(1, min(5, int(a.get("note", 5)))))
            except (TypeError, ValueError):
                notes.append(5)
        business["aggregateRating"] = {
            "@type": "AggregateRating",
            "ratingValue": round(sum(notes) / len(notes), 1),
            "reviewCount": len(notes),
            "bestRating": 5,
        }
    schemas.append(business)

    if faq_pairs:
        schemas.append({
            "@context": "https://schema.org",
            "@type": "FAQPage",
            "mainEntity": [
                {
                    "@type": "Question",
                    "name": q,
                    "acceptedAnswer": {"@type": "Answer", "text": a},
                }
                for q, a in faq_pairs
            ],
        })

    if breadcrumb:
        schemas.append({
            "@context": "https://schema.org",
            "@type": "BreadcrumbList",
            "itemListElement": [
                {
                    "@type": "ListItem",
                    "position": i + 1,
                    "name": name,
                    "item": f"{base}{path}",
                }
                for i, (name, path) in enumerate(breadcrumb)
            ],
        })

    return "\n".join(
        '<script type="application/ld+json">'
        + json.dumps(s, ensure_ascii=False)
        + "</script>"
        for s in schemas
    )


def layout_seed_from(domain: str) -> int:
    """Graine déterministe par domaine — chaque site du réseau a son propre
    ordre de sections (anti-footprint), stable d'un export à l'autre."""
    import zlib
    return zlib.crc32(domain.strip().lower().encode())


def scrape_page(url: str, timeout: float = 15.0) -> dict:
    """Telecharge une URL et en extrait titre, H1, H2s et contenu texte."""
    headers = {
        "User-Agent": (
            "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
            "AppleWebKit/537.36 (KHTML, like Gecko) "
            "Chrome/120.0.0.0 Safari/537.36"
        ),
    }
    try:
        with httpx.Client(follow_redirects=True, timeout=timeout) as client:
            resp = client.get(url, headers=headers)
            resp.raise_for_status()
    except (httpx.HTTPError, httpx.TimeoutException) as exc:
        logger.warning("Echec du scraping de %s : %s", url, exc)
        return {"url": url, "error": str(exc), "title": "", "h1": "", "h2s": [], "content": ""}

    soup = BeautifulSoup(resp.text, "html.parser")
    title_tag = soup.find("title")
    h1_tag = soup.find("h1")
    h2_tags = soup.find_all("h2")

    for tag in soup(["script", "style", "nav", "footer", "header", "aside"]):
        tag.decompose()
    body_text = soup.get_text(separator="\n", strip=True)[:5000]

    return {
        "url": url,
        "title": title_tag.get_text(strip=True) if title_tag else "",
        "h1": h1_tag.get_text(strip=True) if h1_tag else "",
        "h2s": [h2.get_text(strip=True) for h2 in h2_tags],
        "content": body_text,
    }


# ---------------------------------------------------------------------------
# Moteur principal
# ---------------------------------------------------------------------------


class GeolocEngine:
    """Orchestre la generation de pages geolocalisees SEO via Claude."""

    DEFAULT_MODEL = "claude-sonnet-4-6"

    def __init__(self, model: Optional[str] = None):
        settings = get_settings()
        self.client = anthropic.Anthropic(api_key=settings.anthropic_api_key)
        self.model = model or self.DEFAULT_MODEL

    # ------------------------------------------------------------------
    # Appels Claude
    # ------------------------------------------------------------------

    def _call_claude(
        self,
        system: str,
        user_prompt: str,
        max_tokens: int = 4096,
        temperature: float = 0.7,
    ) -> str:
        """Envoie un prompt a Claude et retourne la reponse texte."""
        try:
            response = self.client.messages.create(
                model=self.model,
                max_tokens=max_tokens,
                temperature=temperature,
                system=system,
                messages=[{"role": "user", "content": user_prompt}],
            )
            return response.content[0].text
        except anthropic.APIError as exc:
            logger.error("Erreur API Claude : %s", exc)
            raise RuntimeError(f"Erreur API Claude : {exc}") from exc

    def _call_claude_json(
        self,
        system: str,
        user_prompt: str,
        max_tokens: int = 4096,
        temperature: float = 0.5,
    ) -> dict | list:
        """Envoie un prompt a Claude et parse la reponse en JSON."""
        raw = self._call_claude(system, user_prompt, max_tokens, temperature)

        # Tenter d'extraire le JSON depuis un bloc ```json ... ```
        match = re.search(r"```(?:json)?\s*([\[{][\s\S]*?[}\]])\s*```", raw)
        if match:
            json_str = match.group(1).strip()
        else:
            # Recherche de la premiere accolade / crochet ouvrant
            json_str = raw.strip()
            for i, ch in enumerate(json_str):
                if ch in ("{", "["):
                    json_str = json_str[i:]
                    break

        try:
            return json.loads(json_str)
        except json.JSONDecodeError as exc:
            raise ValueError(
                f"Impossible de parser le JSON retourne par Claude : {exc}\n"
                f"Reponse brute (500 premiers caracteres) : {raw[:500]}"
            ) from exc

    # ------------------------------------------------------------------
    # 1. Analyse du texte source
    # ------------------------------------------------------------------

    def analyze_source(self, source_text: str, brief: dict) -> dict:
        """
        Analyse un texte source et un brief pour en extraire le contexte SEO.

        Args:
            source_text: texte brut (depuis DOCX ou colle).
            brief: dict avec des indications (ex: {"secteur": "plomberie"}).

        Returns:
            dict avec les cles : keyword, slug_base, secteur, template_title,
            template_h1, template_meta, services_cles, contexte_metier.
        """
        system = (
            "Tu es un expert SEO francophone specialise dans le contenu geolocalisé. "
            "Tu analyses des textes sources pour en extraire la structure semantique. "
            "Tu reponds exclusivement en JSON valide."
        )
        user_prompt = f"""Analyse le texte source ci-dessous et le brief associe.
Extrais les informations suivantes en JSON :

- "keyword" : mot-cle principal (2 a 4 mots, sans nom de ville)
- "slug_base" : slug URL de base (sans ville), ex: "plombier"
- "secteur" : description courte du secteur d'activite
- "template_title" : titre SEO avec le placeholder __VILLE__, max 60 caracteres
- "template_h1" : titre H1 avec __VILLE__
- "template_meta" : meta description avec __VILLE__, max 155 caracteres
- "services_cles" : liste de 5 a 8 services cles mentionnes
- "contexte_metier" : resume du contexte metier en 2-3 phrases

BRIEF :
{json.dumps(brief, ensure_ascii=False)}

TEXTE SOURCE :
{source_text[:6000]}

Reponds uniquement avec le JSON, sans commentaire."""

        return self._call_claude_json(system, user_prompt)

    # ------------------------------------------------------------------
    # 2. Generation des ancres de maillage interne
    # ------------------------------------------------------------------

    def generate_anchors(self, ctx: dict) -> list[str]:
        """
        Genere 6 ancres de liens internes naturelles avec le placeholder __VILLE__.
        """
        system = (
            "Tu es un expert en maillage interne SEO. "
            "Tu generes des textes d'ancre naturels et varies. "
            "Reponds en JSON (liste de strings)."
        )
        user_prompt = f"""Genere exactement 6 textes d'ancre pour des liens internes.
Chaque ancre doit :
- contenir le placeholder __VILLE__
- etre naturelle et variee (pas de repetition de formulation)
- integrer le secteur : {ctx.get("secteur", "")}
- etre courte (3 a 8 mots)

Mot-cle principal : {ctx.get("keyword", "")}
Services cles : {", ".join(ctx.get("services_cles", []))}

Reponds avec un tableau JSON de 6 strings, rien d'autre."""

        raw = self._call_claude_json(system, user_prompt)
        if isinstance(raw, list):
            return raw
        # Claude peut renvoyer un dict contenant une liste
        for val in raw.values():
            if isinstance(val, list):
                return val
        return []

    # ------------------------------------------------------------------
    # 3. Decoupage et optimisation en blocs
    # ------------------------------------------------------------------

    def optimize_blocks(
        self,
        source_text: str,
        ctx: dict,
        anchors: list[str],
        paa_questions: Optional[list[str]] = None,
    ) -> dict:
        """
        Decoupe et optimise le contenu source en 6 blocs :
        BLOC_INTRO, BLOC_H2_1 .. BLOC_H2_4, BLOC_FAQ (+ BLOC_AVIS).

        Si *paa_questions* est fourni (People Also Ask réels de Google),
        la FAQ reprend ces questions mot pour mot — meilleure chance de
        featured snippet / position 0.
        """
        system = (
            "Tu es un redacteur SEO expert en contenu geolocalisé. "
            "Tu structures du contenu en blocs HTML optimises. "
            "Utilise __VILLE__ comme placeholder pour le nom de ville. "
            "Reponds exclusivement en JSON valide."
        )
        anchors_str = "\n".join(f"- {a}" for a in anchors)
        paa_block = ""
        if paa_questions:
            paa_block = (
                "\n\nQUESTIONS RÉELLES GOOGLE (People Also Ask) — le BLOC_FAQ doit "
                "reprendre EXACTEMENT ces questions comme H3 (mot pour mot), avec des "
                "réponses concises de 40-60 mots optimisées featured snippet :\n"
                + "\n".join(f"- {q}" for q in paa_questions[:8])
            )

        user_prompt = f"""A partir du texte source et du contexte ci-dessous, genere 6 blocs de contenu HTML.

CONTEXTE :
- Mot-cle : {ctx.get("keyword", "")}
- Secteur : {ctx.get("secteur", "")}
- Services : {", ".join(ctx.get("services_cles", []))}
- Contexte metier : {ctx.get("contexte_metier", "")}

ANCRES DE MAILLAGE (a integrer naturellement, 1 ou 2 par bloc) :
{anchors_str}{paa_block}

TEXTE SOURCE :
{source_text[:6000]}

STRUCTURE ATTENDUE (JSON) :
{{
  "BLOC_INTRO": "<p>Introduction engageante sans H2, avec __VILLE__. 100-150 mots.</p>",
  "BLOC_H2_1": "<h2>Titre section 1 a __VILLE__</h2><h3>...</h3><p>...</p>...",
  "BLOC_H2_2": "<h2>Titre section 2 a __VILLE__</h2><h3>...</h3><p>...</p>...",
  "BLOC_H2_3": "<h2>Titre section 3 a __VILLE__</h2><h3>...</h3><p>...</p>...",
  "BLOC_H2_4": "<h2>Titre section 4 a __VILLE__</h2><h3>...</h3><p>...</p>...",
  "BLOC_FAQ": "<h2>Questions frequentes sur ... a __VILLE__</h2><h3>Q1 ?</h3><p>R1</p>...",
  "BLOC_AVIS": [{{"note": 5, "texte": "Avis client realiste de 2-3 phrases, concret, humanise."}}, ...]
}}

CONSIGNES :
- Chaque BLOC_H2 : 1 H2 + 2-3 H3 + paragraphes riches (200-300 mots)
- BLOC_FAQ : 5 questions/reponses avec H3 pour chaque question
- BLOC_AVIS : 10 a 12 avis clients realistes et varies (notes 4 ou 5), textes concrets lies au metier, sans nom de ville (pas de __VILLE__ dans les avis)
- Ancres sous forme de <a href="__LIEN__">ancre</a>
- __VILLE__ present dans chaque bloc au moins une fois
- Mets en gras avec <strong> le mot-cle principal (2-3 occurrences par bloc, jamais dans les H2/H3)
- Mets en gras avec <strong> les variantes semantiques importantes (1 occurrence chacune)
- HTML propre : h2, h3, p, ul, li, a, strong — pas de markdown
- Contenu expert avec des donnees concretes

Reponds uniquement avec le JSON."""

        return self._call_claude_json(system, user_prompt, max_tokens=8192)

    # ------------------------------------------------------------------
    # 4. Generation de variantes de contenu
    # ------------------------------------------------------------------

    def generate_variants(self, blocs: dict, ctx: dict, count: int = 3) -> list[dict]:
        """
        Genere *count* variantes des blocs pour eviter le duplicate content.

        Chaque variante conserve la meme structure HTML mais reformule
        le texte (synonymes, tournures, exemples differents).
        """
        variants: list[dict] = []
        system = (
            "Tu es un redacteur SEO expert. Tu reecris du contenu HTML "
            "en conservant la structure (memes balises H2, H3) mais en variant "
            "les formulations, exemples et tournures de phrases. "
            "Le contenu doit rester naturel, expert et optimise SEO. "
            "Garde le placeholder __VILLE__. Reponds en JSON valide."
        )

        for i in range(count):
            user_prompt = f"""Reecris les blocs de contenu ci-dessous pour creer la variante {i + 1}/{count}.

CONTEXTE :
- Mot-cle : {ctx.get("keyword", "")}
- Secteur : {ctx.get("secteur", "")}

BLOCS ORIGINAUX :
{json.dumps(blocs, ensure_ascii=False, indent=2)}

CONSIGNES :
- Conserve exactement la meme structure HTML (memes H2, H3)
- Modifie au moins 60 % du texte : reformulations, synonymes, exemples differents
- Garde les ancres de liens (<a>) intactes
- Conserve les mises en gras <strong> du mot-cle principal et de ses variantes (repositionne-les naturellement)
- Garde __VILLE__ comme placeholder
- La variante {i + 1} doit etre significativement differente des precedentes
- Retourne le JSON avec les memes cles (BLOC_INTRO, BLOC_H2_1 .. BLOC_H2_4, BLOC_FAQ)

Reponds uniquement avec le JSON."""

            try:
                variant = self._call_claude_json(system, user_prompt, max_tokens=8192)
                variants.append(variant)
            except (ValueError, RuntimeError) as exc:
                logger.error("Erreur generation variante %d : %s", i + 1, exc)
                variants.append(blocs)  # fallback : reutiliser l'original

        return variants

    # ------------------------------------------------------------------
    # 5. Analyse concurrentielle
    # ------------------------------------------------------------------

    def analyze_competitors(self, urls: list[str], ctx: dict) -> dict:
        """
        Scrape les pages concurrentes et les analyse via Claude.

        Returns:
            dict avec themes_manquants, arguments_cles, enrichissement.
        """
        scraped: list[dict] = []
        for url in urls[:5]:
            data = scrape_page(url)
            if not data.get("error"):
                scraped.append(data)

        if not scraped:
            return {
                "themes_manquants": [],
                "arguments_cles": [],
                "enrichissement": "Aucune page concurrente n'a pu etre analysee.",
            }

        competitors_text = ""
        for s in scraped:
            competitors_text += (
                f"\n--- {s['url']} ---\n"
                f"Title : {s['title']}\n"
                f"H1 : {s['h1']}\n"
                f"H2s : {', '.join(s['h2s'][:10])}\n"
                f"Extrait : {s['content'][:1500]}\n"
            )

        system = (
            "Tu es un expert SEO qui analyse la concurrence. "
            "Reponds en JSON valide."
        )
        user_prompt = f"""Analyse les pages concurrentes ci-dessous par rapport a notre contexte.

NOTRE CONTEXTE :
- Mot-cle : {ctx.get("keyword", "")}
- Secteur : {ctx.get("secteur", "")}
- Services : {", ".join(ctx.get("services_cles", []))}

PAGES CONCURRENTES :
{competitors_text}

Retourne un JSON avec :
- "themes_manquants" : liste de themes que les concurrents couvrent et qui manquent dans notre contenu
- "arguments_cles" : liste d'arguments commerciaux ou techniques pertinents reperes
- "enrichissement" : conseils en 3-5 phrases pour enrichir notre contenu

Reponds uniquement avec le JSON."""

        return self._call_claude_json(system, user_prompt)

    # ------------------------------------------------------------------
    # 6. Maillage interne (liens vers villes proches)
    # ------------------------------------------------------------------

    def generate_maillage(
        self,
        ville: str,
        all_villes: list[dict],
        slug_base: str,
        anchors: list[str],
        nb: int = 8,
    ) -> str:
        """
        Genere un bloc HTML ``<ul>`` de liens internes vers les *nb* villes
        les plus proches geographiquement de *ville*.

        Chaque element de *all_villes* doit avoir les cles :
            nom, slug, lat, lon.
        """
        # Identifier la ville courante
        current = None
        for v in all_villes:
            if v.get("nom", "").lower() == ville.lower():
                current = v
                break

        if not current or "lat" not in current or "lon" not in current:
            return ""

        # Calculer les distances et trier par proximite
        distances: list[tuple[dict, float]] = []
        for v in all_villes:
            if v.get("nom", "").lower() == ville.lower():
                continue
            if "lat" not in v or "lon" not in v:
                continue
            dist = haversine(current["lat"], current["lon"], v["lat"], v["lon"])
            distances.append((v, dist))

        distances.sort(key=lambda x: x[1])
        closest = distances[:nb]
        if not closest:
            return ""

        lines = ['<ul class="maillage-interne">']
        for i, (v, _dist) in enumerate(closest):
            anchor_tpl = (
                anchors[i % len(anchors)] if anchors else "Nos services a __VILLE__"
            )
            anchor_text = replace_ville(anchor_tpl, v["nom"])
            v_slug = v.get("slug", slugify(v["nom"]))
            href = f"/{slug_base}-{v_slug}/"
            lines.append(
                f'  <li><a href="{href}" title="{anchor_text}">{anchor_text}</a></li>'
            )
        lines.append("</ul>")
        return "\n".join(lines)

    # ------------------------------------------------------------------
    # 7. Assemblage de la page complete
    # ------------------------------------------------------------------

    @staticmethod
    def render_avis(avis: list, start: int = 0, count: int = 6) -> list[dict]:
        """Sélectionne *count* avis en rotation à partir de l'index *start*.

        La rotation par ville évite le duplicate content entre les pages.
        """
        clean = [
            a for a in (avis or [])
            if isinstance(a, dict) and a.get("texte")
        ]
        if not clean:
            return []
        return [clean[(start + i) % len(clean)] for i in range(min(count, len(clean)))]

    @staticmethod
    def _stars(note) -> str:
        try:
            n = max(1, min(5, int(note)))
        except (TypeError, ValueError):
            n = 5
        return "★" * n + "☆" * (5 - n)

    def assemble_page(
        self,
        blocs: dict,
        ville: dict,
        ctx: dict,
        maillage: str,
        variant_idx: int = 0,
        use_divi: bool = False,
        images: Optional[list[str]] = None,
        colors: Optional[dict] = None,
        h1: str = "",
        video_url: str = "",
        avis: Optional[list] = None,
        layout_seed: int = 0,
        jsonld: str = "",
        extra_html: str = "",
    ) -> str:
        """
        Assemble les blocs, le maillage, les images, les avis clients et la
        vidéo en une page HTML complete (ou en shortcodes Divi si *use_divi*).

        *layout_seed* fait varier l'ordre des sections (anti-footprint réseau),
        *jsonld* est ajouté en fin de contenu (Schema.org), *extra_html* permet
        d'injecter une section supplémentaire (ex: annuaire des villes d'un
        département sur les pages hub).

        *ville* doit contenir au minimum ``{"nom": "Paris"}``.
        """
        nom_ville = ville.get("nom", "")
        images = images or []
        colors = colors or {"primary": "#2EA3F2", "secondary": "#E02B20"}

        # Remplacement des placeholders dans chaque bloc texte
        page_blocs: dict[str, str] = {}
        for key, html in blocs.items():
            if key == "BLOC_AVIS":
                continue  # liste d'avis, gérée séparément
            page_blocs[key] = replace_ville(str(html), nom_ville)

        # Remplacement du placeholder de lien __LIEN__
        slug_base = ctx.get("slug_base", "service")
        ville_slug = ville.get("slug", slugify(nom_ville))
        for key in page_blocs:
            page_blocs[key] = page_blocs[key].replace(
                "__LIEN__", f"/{slug_base}-{ville_slug}/"
            )

        # Avis en rotation (décalés par ville pour l'unicité)
        avis_list = avis if avis is not None else blocs.get("BLOC_AVIS", [])
        selected_avis = self.render_avis(avis_list, start=variant_idx, count=6)

        keyword = ctx.get("keyword", "")
        zone_label = ville.get("zone_label") or nom_ville

        if use_divi:
            content = self._wrap_divi(
                page_blocs, maillage, images, colors,
                h1=h1, video_url=video_url, avis=selected_avis,
                keyword=keyword, zone_label=zone_label,
                layout_seed=layout_seed, extra_html=extra_html,
            )
        else:
            content = self._wrap_html(
                page_blocs, maillage, images,
                h1=h1, video_url=video_url, avis=selected_avis,
                keyword=keyword, zone_label=zone_label,
                layout_seed=layout_seed, extra_html=extra_html,
            )
        if jsonld:
            content += "\n" + jsonld
        return content

    # Ordres de sections possibles (déterminés par layout_seed) : chaque site
    # du réseau assemble FAQ / maillage / avis / vidéo dans un ordre différent.
    TAIL_ORDERS = [
        ("faq", "maillage", "avis", "video"),
        ("faq", "avis", "maillage", "video"),
        ("avis", "faq", "maillage", "video"),
        ("faq", "video", "avis", "maillage"),
        ("maillage", "faq", "avis", "video"),
        ("avis", "video", "faq", "maillage"),
    ]

    # --- assemblage HTML simple -------------------------------------------

    def _avis_html(self, avis: list[dict]) -> str:
        if not avis:
            return ""
        notes = []
        for a in avis:
            try:
                notes.append(max(1, min(5, int(a.get("note", 5)))))
            except (TypeError, ValueError):
                notes.append(5)
        moyenne = round(sum(notes) / len(notes), 1)
        parts = [
            '<div class="avis-clients">',
            "<h2>Ce que disent nos clients</h2>",
            f"<p style='color:#F1C40F;font-size:22px'>★★★★★</p>"
            f"<p style='color:#666;font-size:14px'>Note {moyenne}/5 — {len(avis)} avis vérifiés</p>",
        ]
        for a, n in zip(avis, notes):
            parts.append(
                f"<div class='avis'>"
                f"<p style='color:#F1C40F;font-size:18px;margin-bottom:8px'>{self._stars(n)}</p>"
                f"<p><em>&laquo;{a.get('texte', '')}&raquo;</em></p>"
                f"</div>"
            )
        parts.append("</div>")
        return "\n".join(parts)

    def _wrap_html(
        self,
        blocs: dict,
        maillage: str,
        images: list[str],
        h1: str = "",
        video_url: str = "",
        avis: Optional[list[dict]] = None,
        keyword: str = "",
        zone_label: str = "",
        layout_seed: int = 0,
        extra_html: str = "",
    ) -> str:
        parts: list[str] = []

        if h1:
            parts.append(f"<h1>{h1}</h1>")

        parts.append(blocs.get("BLOC_INTRO", ""))

        if images:
            parts.append(
                f'<figure><img src="{images[0]}" alt="{keyword} {zone_label}" loading="lazy" /></figure>'
            )

        for i in range(1, 5):
            key = f"BLOC_H2_{i}"
            if key in blocs:
                parts.append(blocs[key])
            if i < len(images):
                parts.append(
                    f'<figure><img src="{images[i]}" alt="{keyword} {zone_label}" loading="lazy" /></figure>'
                )

        if extra_html:
            parts.append(extra_html)

        # Sections de fin dans un ordre propre à chaque site (layout_seed)
        tail: dict[str, str] = {}
        if "BLOC_FAQ" in blocs:
            tail["faq"] = blocs["BLOC_FAQ"]
        if maillage:
            tail["maillage"] = (
                '<div class="maillage-section">\n'
                f"<p style='text-align:center;font-weight:700;font-size:16px;margin-bottom:14px'>"
                f"Retrouvez <strong>{keyword}</strong> autour de {zone_label}</p>\n"
                f"{maillage}\n</div>"
            )
        avis_html = self._avis_html(avis or [])
        if avis_html:
            tail["avis"] = avis_html
        if video_url:
            tail["video"] = (
                f'<div class="video-section" style="text-align:center">'
                f'<iframe width="560" height="315" src="{video_url}" frameborder="0" '
                f'allowfullscreen loading="lazy"></iframe></div>'
            )
        for section in self.TAIL_ORDERS[layout_seed % len(self.TAIL_ORDERS)]:
            if section in tail:
                parts.append(tail[section])

        return "\n\n".join(p for p in parts if p)

    # --- assemblage Divi Builder ------------------------------------------

    def _wrap_divi(
        self,
        blocs: dict,
        maillage: str,
        images: list[str],
        colors: dict,
        h1: str = "",
        video_url: str = "",
        avis: Optional[list[dict]] = None,
        keyword: str = "",
        zone_label: str = "",
        layout_seed: int = 0,
        extra_html: str = "",
    ) -> str:
        """Layout Divi riche, calqué sur le fichier d'import de référence :
        H1 stylé, sections texte/image alternées (2 colonnes), FAQ, maillage
        titré, avis clients en cartes, vidéo, CTA final. L'ordre des sections
        de fin varie selon *layout_seed* (anti-footprint).
        """
        primary = colors.get("primary", "#2EA3F2")
        secondary = colors.get("secondary", "#E02B20")

        TEXT_STYLE = (
            "text_font='||||||||' text_font_size='16px' text_line_height='1.9em' "
            f"header_2_font='|700|||||||' header_2_font_size='26px' header_2_text_color='{primary}' "
            f"header_3_font='|600|||||||' header_3_text_color='#333333'"
        )
        IMG_STYLE = (
            "align='center' max_width='90%' border_radii='on|8px|8px|8px|8px' "
            "box_shadow_style='preset1' box_shadow_blur='30px' "
            "box_shadow_color='rgba(0,0,0,0.12)'"
        )

        def text_section(content: str, bg: str = "#FFFFFF") -> str:
            return (
                f"[et_pb_section _builder_version='4.27.0' background_color='{bg}' "
                f"custom_padding='30px|0px|30px|0px|false|false']"
                f"[et_pb_row _builder_version='4.27.0' max_width='960px']"
                f"[et_pb_column type='4_4' _builder_version='4.27.0']"
                f"[et_pb_text _builder_version='4.27.0' {TEXT_STYLE}]{content}[/et_pb_text]"
                f"[/et_pb_column][/et_pb_row][/et_pb_section]"
            )

        def two_col_section(content: str, img: str, img_left: bool, bg: str) -> str:
            text_col = (
                f"[et_pb_column type='1_2' _builder_version='4.27.0']"
                f"[et_pb_text _builder_version='4.27.0' {TEXT_STYLE}]{content}[/et_pb_text]"
                f"[/et_pb_column]"
            )
            img_col = (
                f"[et_pb_column type='1_2' _builder_version='4.27.0' custom_padding='40px||||false|false']"
                f"[et_pb_image src='{img}' alt='{keyword} {zone_label}' _builder_version='4.27.0' {IMG_STYLE}][/et_pb_image]"
                f"[/et_pb_column]"
            )
            cols = img_col + text_col if img_left else text_col + img_col
            return (
                f"[et_pb_section _builder_version='4.27.0' background_color='{bg}' "
                f"custom_padding='30px|0px|30px|0px|false|false']"
                f"[et_pb_row column_structure='1_2,1_2' _builder_version='4.27.0' "
                f"use_custom_gutter='on' gutter_width='3' equalize_columns='on']"
                f"{cols}[/et_pb_row][/et_pb_section]"
            )

        parts: list[str] = []

        # En-tête H1
        if h1:
            parts.append(
                f"[et_pb_section _builder_version='4.27.0' background_color='#FFFFFF' "
                f"custom_padding='30px|0px|10px|0px|false|false']"
                f"[et_pb_row _builder_version='4.27.0' max_width='960px']"
                f"[et_pb_column type='4_4' _builder_version='4.27.0']"
                f"[et_pb_text _builder_version='4.27.0' header_font='|700|||||||' "
                f"header_font_size='32px' header_text_color='#222222' header_line_height='1.3em' "
                f"custom_padding='||10px||false|false']<h1>{h1}</h1>[/et_pb_text]"
                f"[/et_pb_column][/et_pb_row][/et_pb_section]"
            )

        # Intro pleine largeur
        parts.append(text_section(blocs.get("BLOC_INTRO", "")))

        # Blocs H2 : alternance texte/image en 2 colonnes quand une image existe.
        # La parité du côté image dépend du layout_seed (anti-footprint).
        bg_alt = ["#FFFFFF", "#F7F9FB", "#FFFFFF", "#F7F9FB"]
        side_flip = (layout_seed >> 3) & 1
        for i in range(1, 5):
            key = f"BLOC_H2_{i}"
            if key not in blocs:
                continue
            img = images[(i - 1) % len(images)] if images else ""
            if img:
                parts.append(two_col_section(blocs[key], img, img_left=((i + side_flip) % 2 == 0), bg=bg_alt[i - 1]))
            else:
                parts.append(text_section(blocs[key], bg_alt[i - 1]))

        # Section supplémentaire (ex: annuaire des villes sur les pages département)
        if extra_html:
            parts.append(text_section(extra_html, "#FFFFFF"))

        # Sections de fin dans un ordre propre à chaque site (layout_seed)
        tail: dict[str, str] = {}

        # FAQ
        if "BLOC_FAQ" in blocs:
            tail["faq"] = text_section(blocs["BLOC_FAQ"], "#F0F4F8")

        # Maillage interne titré
        if maillage:
            titre = (
                f"<p style='text-align:center;font-weight:700;font-size:16px;margin-bottom:14px'>"
                f"Retrouvez <strong>{keyword}</strong> autour de {zone_label}</p>"
            )
            tail["maillage"] = text_section(f"{titre}\n{maillage}", "#EAF2E3")

        # Avis clients en cartes (3 colonnes max par rangée)
        if avis:
            notes = []
            for a in avis:
                try:
                    notes.append(max(1, min(5, int(a.get("note", 5)))))
                except (TypeError, ValueError):
                    notes.append(5)
            moyenne = round(sum(notes) / len(notes), 1)
            header = (
                f"[et_pb_row _builder_version='4.27.0' max_width='960px']"
                f"[et_pb_column type='4_4' _builder_version='4.27.0']"
                f"[et_pb_text _builder_version='4.27.0' text_orientation='center' "
                f"header_2_font='|700|||||||' header_2_font_size='26px' header_2_text_color='{primary}']"
                f"<h2>Ce que disent nos clients</h2>"
                f"<p style='color:#F1C40F;font-size:22px'>★★★★★</p>"
                f"<p style='color:#666;font-size:14px'>Note {moyenne}/5 — {len(avis)} avis vérifiés</p>"
                f"[/et_pb_text][/et_pb_column][/et_pb_row]"
            )
            cards_rows = []
            for chunk_start in range(0, len(avis), 3):
                chunk = list(zip(avis, notes))[chunk_start:chunk_start + 3]
                cols = "".join(
                    f"[et_pb_column type='1_3' _builder_version='4.27.0']"
                    f"[et_pb_text _builder_version='4.27.0' background_color='#FFFFFF' "
                    f"border_radii='on|8px|8px|8px|8px' custom_padding='18px|18px|18px|18px|true|true' "
                    f"text_font_size='14px' text_line_height='1.7em' text_text_color='#444444']"
                    f"<p style='color:#F1C40F;font-size:18px;margin-bottom:8px'>{self._stars(n)}</p>"
                    f"<p><em>&laquo;{a.get('texte', '')}&raquo;</em></p>"
                    f"[/et_pb_text][/et_pb_column]"
                    for a, n in chunk
                )
                cards_rows.append(
                    f"[et_pb_row column_structure='1_3,1_3,1_3' _builder_version='4.27.0' "
                    f"max_width='960px' column_padding_mobile='on']{cols}[/et_pb_row]"
                )
            tail["avis"] = (
                f"[et_pb_section _builder_version='4.27.0' background_color='#F7F9FB' "
                f"custom_padding='40px|0px|40px|0px|false|false']{header}{''.join(cards_rows)}[/et_pb_section]"
            )

        # Vidéo
        if video_url:
            tail["video"] = (
                f"[et_pb_section _builder_version='4.27.0' background_color='#FFFFFF' "
                f"custom_padding='30px|0px|30px|0px|false|false']"
                f"[et_pb_row _builder_version='4.27.0' max_width='960px']"
                f"[et_pb_column type='4_4' _builder_version='4.27.0']"
                f"[et_pb_video src='{video_url}' _builder_version='4.27.0' "
                f"max_width='720px' module_alignment='center'][/et_pb_video]"
                f"[/et_pb_column][/et_pb_row][/et_pb_section]"
            )

        for section in self.TAIL_ORDERS[layout_seed % len(self.TAIL_ORDERS)]:
            if section in tail:
                parts.append(tail[section])

        # CTA final
        cta = (
            f'[et_pb_section fb_built="1" background_color="{primary}" '
            f'custom_padding="30px||30px||false|false"]'
            f'[et_pb_row][et_pb_column type="4_4"]'
            f'[et_pb_cta title="Contactez-nous" '
            f'button_text="Demander un devis" '
            f'button_url="#contact" '
            f'use_background_color="off" '
            f'header_text_color="#ffffff" '
            f'body_text_color="#ffffff" '
            f'custom_button="on" '
            f'button_text_color="#ffffff" '
            f'button_bg_color="{secondary}" '
            f'button_border_color="{secondary}"]'
            f"<p>Nous intervenons rapidement. Contactez notre equipe "
            f"pour un devis gratuit.</p>"
            f"[/et_pb_cta][/et_pb_column][/et_pb_row][/et_pb_section]"
        )
        parts.append(cta)

        return "\n".join(parts)

    # ------------------------------------------------------------------
    # 8. Generation du fichier Excel
    # ------------------------------------------------------------------

    def build_excel(
        self,
        villes: list[dict],
        ctx: dict,
        blocs: dict,
        variants: list[dict],
        config: dict,
    ) -> bytes:
        """
        Genere un fichier Excel d'import WordPress (WP All Import).

        Colonnes produites :
            Ville actuelle, SLUG, post_title, H1, post_description,
            post_content, post_thumbnail, post_date, post_author,
            post_category, post_tag, post_status, Population, Type,
            _et_pb_use_builder, _et_pb_old_content

        *config* attend les cles :
            post_author, post_category, post_tag, post_status,
            post_date_start (YYYY-MM-DD), use_divi, images, colors, anchors.
        """
        wb = Workbook()
        ws = wb.active
        ws.title = "Pages Geoloc"

        headers = [
            "Ville actuelle",
            "SLUG",
            "post_title",
            "H1",
            "post_description",
            "post_content",
            "post_thumbnail",
            "post_date",
            "post_author",
            "post_category",
            "post_tag",
            "post_status",
            "Population",
            "Type",
            "_et_pb_use_builder",
            "_et_pb_old_content",
        ]

        header_font = Font(bold=True, color="FFFFFF", size=11)
        header_fill = PatternFill(
            start_color="2E86C1", end_color="2E86C1", fill_type="solid"
        )
        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")

        # Lecture de la configuration
        slug_base = ctx.get("slug_base", "service")
        template_title = ctx.get("template_title", "__VILLE__")
        template_h1 = ctx.get("template_h1", "__VILLE__")
        template_meta = ctx.get("template_meta", "__VILLE__")
        anchors = config.get("anchors", [])
        use_divi = config.get("use_divi", False)
        images = config.get("images", [])
        colors = config.get("colors", {"primary": "#2EA3F2", "secondary": "#E02B20"})
        post_author = config.get("post_author", "admin")
        post_category = config.get("post_category", "")
        post_tag = config.get("post_tag", "")
        post_status = config.get("post_status", "draft")
        post_date_start = config.get("post_date_start", "2026-01-15")

        # Rotation original + variantes
        all_blocs = [blocs] + variants

        for row_idx, ville in enumerate(villes, start=2):
            nom = ville.get("nom", "")
            ville_slug = ville.get("slug", slugify(nom))

            variant_idx = (row_idx - 2) % len(all_blocs)
            current_blocs = all_blocs[variant_idx]

            # Maillage interne
            maillage = self.generate_maillage(
                nom, villes, slug_base, anchors, nb=8
            )

            # Assemblage de la page
            content = self.assemble_page(
                current_blocs,
                ville,
                ctx,
                maillage,
                variant_idx=variant_idx,
                use_divi=use_divi,
                images=images,
                colors=colors,
            )

            # Date incrementale (+1 jour par ville)
            try:
                base_date = datetime.strptime(post_date_start, "%Y-%m-%d")
                post_date = (base_date + timedelta(days=row_idx - 2)).strftime(
                    "%Y-%m-%d"
                )
            except (ValueError, TypeError):
                post_date = post_date_start

            slug_full = f"{slug_base}-{ville_slug}"
            title = replace_ville(template_title, nom)
            h1 = replace_ville(template_h1, nom)
            meta = replace_ville(template_meta, nom)

            row_data = [
                nom,                                    # Ville actuelle
                slug_full,                              # SLUG
                title,                                  # post_title
                h1,                                     # H1
                meta,                                   # post_description
                content,                                # post_content
                images[0] if images else "",             # post_thumbnail
                post_date,                              # post_date
                post_author,                            # post_author
                post_category,                          # post_category
                post_tag,                               # post_tag
                post_status,                            # post_status
                ville.get("population", ""),             # Population
                ville.get("type", ""),                   # Type
                "on" if use_divi else "",                # _et_pb_use_builder
                "",                                      # _et_pb_old_content
            ]

            for col_idx, value in enumerate(row_data, start=1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                if col_idx == 6:
                    cell.alignment = Alignment(wrap_text=True)

        # Largeurs de colonnes
        col_widths = {
            1: 18,  2: 30,  3: 50,  4: 50,  5: 60,
            6: 80,  7: 40,  8: 14,  9: 12,  10: 20,
            11: 20, 12: 10, 13: 12, 14: 12, 15: 18, 16: 15,
        }
        for col, width in col_widths.items():
            letter = ws.cell(row=1, column=col).column_letter
            ws.column_dimensions[letter].width = width

        buffer = io.BytesIO()
        wb.save(buffer)
        return buffer.getvalue()
